"""
Sprint 8 — Dashboard & Relatórios
Endpoints de KPIs, resumos, gestão de prazos e exportação Excel.
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from typing import Annotated, Any, Optional

import sqlalchemy as sa
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.deps import CurrentUser, InternalOnly, get_session
from app.models.client import Client, ClientType
from app.models.financial import FinancialEntry
from app.models.meeting import Meeting
from app.models.procedure import PROCEDURE_TYPE_LABELS, Procedure, ProcedureStage
from app.models.quote import Contract, Quote

router = APIRouter()

# ── helpers ───────────────────────────────────────────────────────────────────

def _zero() -> Decimal:
    return Decimal("0.00")


async def _scalar(db: AsyncSession, stmt) -> Any:
    return (await db.execute(stmt)).scalar_one_or_none() or 0


def _client_name(client: Client | None) -> str | None:
    if not client:
        return None
    if client.client_type == ClientType.PF and client.pf_data:
        return client.pf_data.name
    if client.client_type == ClientType.PJ and client.pj_data:
        return client.pj_data.company_name
    return None


def _proc_number(p: Procedure) -> str:
    year = p.opened_at.year if p.opened_at else datetime.now().year
    return f"BMB-{year}-{str(p.protocol_number).zfill(4)}"


def _proc_opts():
    return [
        selectinload(Procedure.client).selectinload(Client.pf_data),
        selectinload(Procedure.client).selectinload(Client.pj_data),
        selectinload(Procedure.responsible),
    ]


# ── Dashboard KPIs ────────────────────────────────────────────────────────────

@router.get("/dashboard")
async def dashboard_kpis(
    _: InternalOnly,
    db: Annotated[AsyncSession, Depends(get_session)],
):
    now = datetime.now(timezone.utc)
    today = now.date()
    first_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    seven_days = today + timedelta(days=7)

    # ── Clientes ──────────────────────────────────────────────────────────────
    total_clientes = await _scalar(
        db, sa.select(sa.func.count()).where(Client.is_active.is_(True))
    )
    clientes_mes = await _scalar(
        db,
        sa.select(sa.func.count()).where(
            Client.is_active.is_(True),
            Client.created_at >= first_of_month,
        ),
    )

    # ── Procedimentos ─────────────────────────────────────────────────────────
    total_proc = await _scalar(db, sa.select(sa.func.count()).select_from(Procedure))
    proc_ativos = await _scalar(
        db, sa.select(sa.func.count()).where(Procedure.status == "em_andamento")
    )
    proc_concluidos_mes = await _scalar(
        db,
        sa.select(sa.func.count()).where(
            Procedure.status == "concluido",
            Procedure.updated_at >= first_of_month,
        ),
    )

    proc_por_tipo_rows = (
        await db.execute(
            sa.select(Procedure.procedure_type, sa.func.count())
            .group_by(Procedure.procedure_type)
            .order_by(sa.func.count().desc())
        )
    ).all()
    proc_por_tipo = [
        {
            "tipo": row[0],
            "label": PROCEDURE_TYPE_LABELS.get(row[0], row[0]),
            "total": row[1],
        }
        for row in proc_por_tipo_rows
    ]

    proc_por_status_rows = (
        await db.execute(
            sa.select(Procedure.status, sa.func.count()).group_by(Procedure.status)
        )
    ).all()
    proc_por_status = {row[0]: row[1] for row in proc_por_status_rows}

    # ── Contratos & Financeiro ────────────────────────────────────────────────
    contratos_ativos = await _scalar(
        db, sa.select(sa.func.count()).where(Contract.status == "assinado")
    )
    orc_assinados = await _scalar(
        db, sa.select(sa.func.count()).where(Quote.status == "assinado")
    )

    contracts_rows = (
        await db.execute(sa.select(Contract).where(Contract.status == "assinado"))
    ).scalars().all()
    hon_a_receber = _zero()
    hon_recebidos_mes = _zero()
    for c in contracts_rows:
        for inst in (c.installments or []):
            v = Decimal(str(inst.get("value", 0)))
            if inst.get("status") == "pendente":
                hon_a_receber += v
            elif inst.get("status") == "pago":
                try:
                    paid_dt = datetime.fromisoformat(inst.get("paid_at", ""))
                    if paid_dt.replace(tzinfo=timezone.utc) >= first_of_month:
                        hon_recebidos_mes += v
                except Exception:
                    pass

    custas_pend_val = await _scalar(
        db,
        sa.select(sa.func.sum(FinancialEntry.value)).where(
            FinancialEntry.tipo == "custa_real",
            FinancialEntry.status == "pendente",
        ),
    )
    repasses_pend_val = await _scalar(
        db,
        sa.select(sa.func.sum(FinancialEntry.value)).where(
            FinancialEntry.tipo == "repasse_despachante",
            FinancialEntry.status == "pendente",
        ),
    )

    # ── Prazos próximos ───────────────────────────────────────────────────────
    etapas_prox = (
        await db.execute(
            sa.select(ProcedureStage)
            .join(Procedure, ProcedureStage.procedure_id == Procedure.id)
            .where(
                ProcedureStage.due_date.isnot(None),
                ProcedureStage.due_date <= seven_days,
                ProcedureStage.due_date >= today,
                ProcedureStage.status != "concluida",
                Procedure.status == "em_andamento",
            )
            .order_by(ProcedureStage.due_date)
            .limit(8)
        )
    ).scalars().all()

    etapas_atraso = (
        await db.execute(
            sa.select(ProcedureStage)
            .join(Procedure, ProcedureStage.procedure_id == Procedure.id)
            .where(
                ProcedureStage.due_date.isnot(None),
                ProcedureStage.due_date < today,
                ProcedureStage.status != "concluida",
                Procedure.status == "em_andamento",
            )
            .order_by(ProcedureStage.due_date)
            .limit(5)
        )
    ).scalars().all()

    async def _enrich_stages(stages):
        result = []
        for s in stages:
            proc = (
                await db.execute(
                    sa.select(Procedure)
                    .where(Procedure.id == s.procedure_id)
                    .options(*_proc_opts())
                )
            ).scalar_one_or_none()
            result.append({
                "stage_id": str(s.id),
                "stage_name": s.name,
                "due_date": s.due_date.isoformat() if s.due_date else None,
                "days_left": (s.due_date - today).days if s.due_date else None,
                "status": s.status,
                "procedure_id": str(s.procedure_id),
                "procedure_number": _proc_number(proc) if proc else "—",
                "client_name": _client_name(proc.client) if proc else None,
            })
        return result

    # ── Reuniões hoje ─────────────────────────────────────────────────────────
    reunioes_hoje = await _scalar(
        db,
        sa.select(sa.func.count()).where(sa.func.date(Meeting.scheduled_at) == today),
    )

    # ── Atividade recente ─────────────────────────────────────────────────────
    procs_recentes = (
        await db.execute(
            sa.select(Procedure)
            .options(*_proc_opts())
            .order_by(Procedure.updated_at.desc())
            .limit(8)
        )
    ).scalars().all()

    return {
        "clientes": {"total": total_clientes, "novos_mes": clientes_mes},
        "procedimentos": {
            "total": total_proc,
            "ativos": proc_ativos,
            "concluidos_mes": proc_concluidos_mes,
            "por_status": proc_por_status,
            "por_tipo": proc_por_tipo[:8],
        },
        "orcamentos": {"assinados": orc_assinados},
        "contratos": {"ativos": contratos_ativos},
        "financeiro": {
            "honorarios_a_receber": float(hon_a_receber),
            "honorarios_recebidos_mes": float(hon_recebidos_mes),
            "custas_pendentes": float(custas_pend_val or 0),
            "repasses_pendentes": float(repasses_pend_val or 0),
        },
        "reunioes_hoje": reunioes_hoje,
        "prazos": {
            "proximos": await _enrich_stages(etapas_prox),
            "em_atraso": await _enrich_stages(etapas_atraso),
        },
        "atividade_recente": [
            {
                "id": str(p.id),
                "numero": _proc_number(p),
                "tipo_label": PROCEDURE_TYPE_LABELS.get(p.procedure_type, p.procedure_type),
                "status": p.status,
                "client_name": _client_name(p.client),
                "updated_at": p.updated_at.isoformat(),
            }
            for p in procs_recentes
        ],
    }


# ── Relatório de procedimentos ────────────────────────────────────────────────

@router.get("/procedures")
async def report_procedures(
    _: InternalOnly,
    db: Annotated[AsyncSession, Depends(get_session)],
    status: Optional[str] = None,
    procedure_type: Optional[str] = None,
):
    q = sa.select(Procedure).options(*_proc_opts()).order_by(Procedure.opened_at.desc())
    if status:
        q = q.where(Procedure.status == status)
    if procedure_type:
        q = q.where(Procedure.procedure_type == procedure_type)

    rows = (await db.execute(q)).scalars().all()

    tipo_stats: dict[str, dict] = {}
    for p in rows:
        t = p.procedure_type
        if t not in tipo_stats:
            tipo_stats[t] = {
                "tipo": t,
                "label": PROCEDURE_TYPE_LABELS.get(t, t),
                "total": 0,
                "ativos": 0,
                "concluidos": 0,
                "cancelados": 0,
            }
        tipo_stats[t]["total"] += 1
        if p.status == "em_andamento":
            tipo_stats[t]["ativos"] += 1
        elif p.status == "concluido":
            tipo_stats[t]["concluidos"] += 1
        elif p.status == "cancelado":
            tipo_stats[t]["cancelados"] += 1

    return {
        "total": len(rows),
        "por_tipo": list(tipo_stats.values()),
        "items": [
            {
                "id": str(p.id),
                "numero": _proc_number(p),
                "tipo": p.procedure_type,
                "tipo_label": PROCEDURE_TYPE_LABELS.get(p.procedure_type, p.procedure_type),
                "status": p.status,
                "client_name": _client_name(p.client),
                "responsible_name": p.responsible.name if p.responsible else None,
                "opened_at": p.opened_at.isoformat() if p.opened_at else None,
                "deadline": p.deadline.isoformat() if p.deadline else None,
            }
            for p in rows
        ],
    }


# ── Relatório financeiro ──────────────────────────────────────────────────────

@router.get("/financial")
async def report_financial(
    _: InternalOnly,
    db: Annotated[AsyncSession, Depends(get_session)],
):
    contracts = (
        await db.execute(sa.select(Contract).where(Contract.status == "assinado"))
    ).scalars().all()

    total_contratado = _zero()
    total_recebido = _zero()
    total_pendente = _zero()
    contrato_items = []

    for c in contracts:
        tv = Decimal(str(c.total_value))
        total_contratado += tv
        pago_c = _zero()
        pend_c = _zero()
        for inst in (c.installments or []):
            v = Decimal(str(inst.get("value", 0)))
            if inst.get("status") == "pago":
                pago_c += v
                total_recebido += v
            else:
                pend_c += v
                total_pendente += v
        contrato_items.append({
            "id": str(c.id),
            "numero": f"BMB-CTR-{c.contract_year}-{str(c.contract_number).zfill(4)}",
            "client_name": c.client_name,
            "total": float(tv),
            "recebido": float(pago_c),
            "pendente": float(pend_c),
            "payment_model": c.payment_model,
        })

    custas = (
        await db.execute(
            sa.select(FinancialEntry).where(
                FinancialEntry.tipo == "custa_real",
                FinancialEntry.status != "cancelado",
            )
        )
    ).scalars().all()
    custas_pagas = sum(Decimal(str(e.value)) for e in custas if e.status == "pago")
    custas_pend = sum(Decimal(str(e.value)) for e in custas if e.status == "pendente")

    repasses = (
        await db.execute(
            sa.select(FinancialEntry).where(
                FinancialEntry.tipo == "repasse_despachante",
                FinancialEntry.status != "cancelado",
            )
        )
    ).scalars().all()
    rep_pagos = sum(Decimal(str(e.value)) for e in repasses if e.status == "pago")
    rep_pend = sum(Decimal(str(e.value)) for e in repasses if e.status == "pendente")

    return {
        "honorarios": {
            "total_contratado": float(total_contratado),
            "total_recebido": float(total_recebido),
            "total_pendente": float(total_pendente),
            "contratos": contrato_items,
        },
        "custas": {
            "pagas": float(custas_pagas),
            "pendentes": float(custas_pend),
            "total": float(custas_pagas + custas_pend),
        },
        "repasses": {
            "pagos": float(rep_pagos),
            "pendentes": float(rep_pend),
            "total": float(rep_pagos + rep_pend),
        },
    }


# ── Gestão de prazos ──────────────────────────────────────────────────────────

@router.get("/deadlines")
async def report_deadlines(
    _: InternalOnly,
    db: Annotated[AsyncSession, Depends(get_session)],
    days_ahead: int = Query(30, ge=1, le=365),
):
    today = date.today()
    until = today + timedelta(days=days_ahead)

    stages = (
        await db.execute(
            sa.select(ProcedureStage)
            .join(Procedure, ProcedureStage.procedure_id == Procedure.id)
            .where(
                ProcedureStage.due_date.isnot(None),
                ProcedureStage.due_date >= today,
                ProcedureStage.due_date <= until,
                ProcedureStage.status != "concluida",
                Procedure.status == "em_andamento",
            )
            .order_by(ProcedureStage.due_date)
        )
    ).scalars().all()

    stage_items = []
    for s in stages:
        proc = (
            await db.execute(
                sa.select(Procedure)
                .where(Procedure.id == s.procedure_id)
                .options(*_proc_opts())
            )
        ).scalar_one_or_none()
        days_left = (s.due_date - today).days if s.due_date else None
        stage_items.append({
            "tipo": "etapa",
            "id": str(s.id),
            "descricao": s.name,
            "due_date": s.due_date.isoformat() if s.due_date else None,
            "days_left": days_left,
            "urgente": days_left is not None and days_left <= 3,
            "procedure_id": str(s.procedure_id),
            "procedure_number": _proc_number(proc) if proc else "—",
            "client_name": _client_name(proc.client) if proc else None,
        })

    entries = (
        await db.execute(
            sa.select(FinancialEntry)
            .where(
                FinancialEntry.due_date.isnot(None),
                FinancialEntry.due_date >= today,
                FinancialEntry.due_date <= until,
                FinancialEntry.status == "pendente",
            )
            .order_by(FinancialEntry.due_date)
        )
    ).scalars().all()

    fin_items = []
    for e in entries:
        days_left = (e.due_date - today).days if e.due_date else None
        fin_items.append({
            "tipo": "financeiro",
            "subtipo": e.tipo,
            "id": str(e.id),
            "descricao": e.description,
            "valor": float(e.value),
            "due_date": e.due_date.isoformat() if e.due_date else None,
            "days_left": days_left,
            "urgente": days_left is not None and days_left <= 3,
            "procedure_id": str(e.procedure_id) if e.procedure_id else None,
        })

    atraso_stages = (
        await db.execute(
            sa.select(ProcedureStage)
            .join(Procedure, ProcedureStage.procedure_id == Procedure.id)
            .where(
                ProcedureStage.due_date.isnot(None),
                ProcedureStage.due_date < today,
                ProcedureStage.status != "concluida",
                Procedure.status == "em_andamento",
            )
            .order_by(ProcedureStage.due_date)
        )
    ).scalars().all()

    atraso_items = []
    for s in atraso_stages:
        proc = (
            await db.execute(
                sa.select(Procedure)
                .where(Procedure.id == s.procedure_id)
                .options(*_proc_opts())
            )
        ).scalar_one_or_none()
        days_late = (today - s.due_date).days if s.due_date else None
        atraso_items.append({
            "tipo": "etapa",
            "id": str(s.id),
            "descricao": s.name,
            "due_date": s.due_date.isoformat() if s.due_date else None,
            "days_late": days_late,
            "procedure_id": str(s.procedure_id),
            "procedure_number": _proc_number(proc) if proc else "—",
            "client_name": _client_name(proc.client) if proc else None,
        })

    all_upcoming = sorted(stage_items + fin_items, key=lambda x: x["due_date"] or "")

    return {
        "days_ahead": days_ahead,
        "proximos": all_upcoming,
        "em_atraso": atraso_items,
        "totais": {
            "proximos": len(all_upcoming),
            "em_atraso": len(atraso_items),
            "urgentes": sum(1 for i in all_upcoming if i.get("urgente")),
        },
    }


# ── Exportação Excel ──────────────────────────────────────────────────────────

@router.get("/export/procedures.xlsx")
async def export_procedures_xlsx(
    _: InternalOnly,
    db: Annotated[AsyncSession, Depends(get_session)],
):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = (
        await db.execute(
            sa.select(Procedure)
            .options(*_proc_opts())
            .order_by(Procedure.opened_at.desc())
        )
    ).scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Procedimentos"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    headers = [
        "Protocolo", "Tipo", "Status", "Cliente", "Responsável",
        "Data Abertura", "Prazo", "Matrícula", "INCRA",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_map = {"em_andamento": "Em andamento", "concluido": "Concluído", "cancelado": "Cancelado"}
    for row_idx, p in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=_proc_number(p))
        ws.cell(row=row_idx, column=2, value=PROCEDURE_TYPE_LABELS.get(p.procedure_type, p.procedure_type))
        ws.cell(row=row_idx, column=3, value=status_map.get(p.status, p.status))
        ws.cell(row=row_idx, column=4, value=_client_name(p.client) or "")
        ws.cell(row=row_idx, column=5, value=p.responsible.name if p.responsible else "")
        ws.cell(row=row_idx, column=6, value=p.opened_at.isoformat() if p.opened_at else "")
        ws.cell(row=row_idx, column=7, value=p.deadline.isoformat() if p.deadline else "")
        ws.cell(row=row_idx, column=8, value=p.matricula or "")
        ws.cell(row=row_idx, column=9, value=p.incra or "")
        if row_idx % 2 == 0:
            fill = PatternFill("solid", fgColor="F0F4F8")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    for i, w in enumerate([18, 30, 16, 28, 20, 16, 16, 18, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=procedimentos.xlsx"},
    )


@router.get("/export/financial.xlsx")
async def export_financial_xlsx(
    _: InternalOnly,
    db: Annotated[AsyncSession, Depends(get_session)],
):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    entries = (
        await db.execute(
            sa.select(FinancialEntry)
            .where(FinancialEntry.status != "cancelado")
            .order_by(FinancialEntry.created_at.desc())
        )
    ).scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lançamentos Financeiros"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    headers = [
        "Nº Repasse", "Tipo", "Categoria", "Descrição",
        "Valor (R$)", "Status", "Vencimento", "Pago em", "Criado em",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    tipo_map = {
        "custa_real": "Custa real",
        "repasse_despachante": "Repasse despachante",
        "honorario_recebido": "Honorário recebido",
    }
    status_map = {"pendente": "Pendente", "pago": "Pago", "cancelado": "Cancelado"}
    cat_map = {
        "cartorio": "Cartório", "imposto": "Imposto", "taxa": "Taxa",
        "diligencia": "Diligência", "despachante": "Despachante",
        "honorario": "Honorário", "outro": "Outro",
    }

    for row_idx, e in enumerate(entries, 2):
        fmt_num = (
            f"BMB-REP-{e.entry_year}-{str(e.entry_number).zfill(4)}"
            if e.tipo == "repasse_despachante" and e.entry_number and e.entry_year
            else ""
        )
        ws.cell(row=row_idx, column=1, value=fmt_num)
        ws.cell(row=row_idx, column=2, value=tipo_map.get(e.tipo, e.tipo))
        ws.cell(row=row_idx, column=3, value=cat_map.get(e.category, e.category))
        ws.cell(row=row_idx, column=4, value=e.description)
        val_cell = ws.cell(row=row_idx, column=5, value=float(e.value))
        val_cell.number_format = 'R$ #,##0.00'
        ws.cell(row=row_idx, column=6, value=status_map.get(e.status, e.status))
        ws.cell(row=row_idx, column=7, value=e.due_date.isoformat() if e.due_date else "")
        ws.cell(row=row_idx, column=8, value=e.paid_at.date().isoformat() if e.paid_at else "")
        ws.cell(row=row_idx, column=9, value=e.created_at.date().isoformat())
        if row_idx % 2 == 0:
            fill = PatternFill("solid", fgColor="F0F4F8")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    for i, w in enumerate([18, 22, 16, 40, 14, 12, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=financeiro.xlsx"},
    )
